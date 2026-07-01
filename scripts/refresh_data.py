"""WFM Intelligence · conversor mensual

Uso local:
  1) Coloca tu Excel como data/base.xlsx
  2) Ejecuta: python scripts/refresh_data.py
  3) Se actualiza data/model.json

Requiere: pip install openpyxl
"""
from __future__ import annotations
import json, re, math
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "data" / "base.xlsx"
OUTPUT = ROOT / "data" / "model.json"
SHEET = "Base_WFM"

DIM_MAP = {
    "region": ["Reg Nom Actual", "Región"],
    "dm": ["DM"],
    "tienda": ["CC Nombre"],
    "cc": ["CC"],
    "mes": ["Fecha mes", "Mes"],
    "mesNum": ["MES NUM"],
    "semana": ["Semana"],
    "formato": ["Tipo de Tienda", "Sub Categoria Ops"],
    "categoria": ["Categoria Ops"],
    "lat": ["LATITUD"],
    "lng": ["LONGITUD"],
}
METRIC_MAP = {
    "IPLH":"IPLH", "TPLH":"TPLH", "Ratio":"Ratio",
    "Exactitud_Kronos":"Exactitud Kronos", "Exactitud_Pronostico_gerente":"Exactitud Pronostico gerente",
    "NC":"% NC", "Avg_Horas_Planificadas":"Avg (Horas Planificadas)",
    "Horas_Fcst":"Horas Fcst", "Horas_Planificadas":"Horas Planificadas", "Horas_Kronos":"Horas Kronos",
    "Emp_Autorizados":"Emp. Autorizados", "Emp_Reales":"Emp. Reales",
    "Items":"Items", "Items_Gerente_Aj":"Items Gerente (Aj.)", "Items_Kronos_Proy":"Items Kronos (Proy.)", "Items_Reales":"Items Reales",
    "Ordenes_Reales":"Ordenes Reales", "Ventas_Reales":"Ventas Reales",
    "Capacidad_Instalada":"Capacidad Instalada", "Gasto_adicional":"Gasto adicional", "Venta_Perdida":"Venta Perdida", "Total_NC":"Total NC",
}

def clean_num(v):
    if v is None or v == "" or str(v).strip() in {"-", "-   ", "-    ", "—"}: return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return 0
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    pct = s.endswith("%")
    s = s.replace("%", "")
    try: n = float(s)
    except ValueError: return 0
    return n

def clean_text(v):
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()

def find_col(headers, names):
    norm = {clean_text(h).lower(): i for i, h in enumerate(headers)}
    for n in names:
        if clean_text(n).lower() in norm: return norm[clean_text(n).lower()]
    return None

def catalog_id(cat, value):
    value = clean_text(value)
    if value not in cat: cat[value] = len(cat)
    return cat[value]

def strip_pago(h):
    s = re.sub(r"^Avg \(NC ", "", clean_text(h)).replace(")", "")
    return s.replace("Adm-", "").replace("Adm.", "Adm").strip()

def strip_nc(h):
    return clean_text(h).replace("NC ", "", 1).strip()

def main():
    if not INPUT.exists():
        raise SystemExit(f"No existe {INPUT}. Coloca el Excel mensual como data/base.xlsx")
    wb = load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [clean_text(x) for x in next(rows_iter)]
    dim_cols = {k: find_col(headers, names) for k, names in DIM_MAP.items()}
    metric_cols = {k: find_col(headers, [h]) for k, h in METRIC_MAP.items()}
    pago_cols = [(i, strip_pago(h)) for i, h in enumerate(headers) if clean_text(h).startswith("Avg (NC ")]
    nc_cols = [(i, strip_nc(h)) for i, h in enumerate(headers) if clean_text(h).startswith("NC ")]
    dims = {k: {} for k in DIM_MAP if k not in {"lat", "lng", "semana", "mesNum"}}
    metric_keys = list(METRIC_MAP.keys())
    pt, nt, pago_cat, nc_cat = [], [], {}, {}
    r, p, nc = [], [], []
    for row_idx, row in enumerate(rows_iter):
        rec=[]
        for k in ["region","dm","tienda","cc","mes"]:
            rec.append(catalog_id(dims[k], row[dim_cols[k]] if dim_cols[k] is not None else ""))
        rec.append(int(clean_num(row[dim_cols["mesNum"]])) if dim_cols["mesNum"] is not None else 0)
        rec.append(int(clean_num(row[dim_cols["semana"]])) if dim_cols["semana"] is not None else 0)
        for k in ["formato","categoria"]:
            rec.append(catalog_id(dims[k], row[dim_cols[k]] if dim_cols[k] is not None else ""))
        rec.append(clean_num(row[dim_cols["lat"]]) if dim_cols["lat"] is not None else 0)
        rec.append(clean_num(row[dim_cols["lng"]]) if dim_cols["lng"] is not None else 0)
        rec.extend(clean_num(row[metric_cols[k]]) if metric_cols[k] is not None else 0 for k in metric_keys)
        r.append(rec)
        rid=len(r)-1
        for ci,name in pago_cols:
            v=clean_num(row[ci])
            if v:
                if name not in pago_cat: pago_cat[name]=len(pt); pt.append(name)
                p.append([rid,pago_cat[name],v])
        for ci,name in nc_cols:
            v=clean_num(row[ci])
            if v:
                if name not in nc_cat: nc_cat[name]=len(nt); nt.append(name)
                nc.append([rid,nc_cat[name],v])
    d_out={k:[None]*len(v) for k,v in dims.items()}
    for k,cat in dims.items():
        for name,i in cat.items(): d_out[k][i]=name
    model={"m":{"rows":len(r),"pago_entries":len(p),"nc_entries":len(nc),"source":INPUT.name,"sheet":ws.title},"d":d_out,"k":metric_keys,"r":r,"pt":pt,"p":p,"nt":nt,"nc":nc}
    OUTPUT.write_text(json.dumps(model, ensure_ascii=False, separators=(",",":")), encoding="utf-8")
    print(f"Modelo actualizado: {OUTPUT} · registros {len(r):,}")

if __name__ == "__main__":
    main()
